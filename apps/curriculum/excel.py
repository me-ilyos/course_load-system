from typing import Dict, List, Optional, Tuple
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from django.core.exceptions import ValidationError
from .schema import CurriculumSchema, CourseType, HourDistribution, SemesterData

class ExcelProcessor:
    """Handles Excel file processing for curriculum data"""

    REQUIRED_COLUMNS = [
        'course_code',
        'course_name',
        'credits',
        'semester',
        'prerequisites'
    ]

    def __init__(self):
        self.schema = CurriculumSchema()

    def validate_headers(self, df: pd.DataFrame) -> None:
        """Validate that all required columns are present"""
        missing_cols = set(self.REQUIRED_COLUMNS) - set(df.columns)
        if missing_cols:
            raise ValidationError(f"Missing required columns: {', '.join(missing_cols)}")

    def clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and prepare the dataframe for processing"""
        # Replace NaN with 0 for numeric columns
        numeric_columns = ['Credits', 'Lecture', 'Lab', 'Practice', 'Seminar', 'Individual']
        df[numeric_columns] = df[numeric_columns].fillna(0)
        
        # Convert numeric columns to integers
        for col in numeric_columns:
            df[col] = df[col].astype(int)
        
        # Clean string columns
        string_columns = ['Course Code', 'Course Name', 'Type']
        for col in string_columns:
            df[col] = df[col].fillna('').str.strip()
        
        return df

    def process_prerequisites(self, prerequisites_str: str) -> List[str]:
        """Process prerequisites string into list of course codes"""
        if not prerequisites_str or pd.isna(prerequisites_str):
            return []
        return [code.strip() for code in prerequisites_str.split(',')]

    def read_excel(self, file_path: str) -> Dict:
        """Read curriculum data from Excel file"""
        try:
            df = pd.read_excel(file_path)
            self.validate_headers(df)
            df = self.clean_dataframe(df)
            
            curriculum_data = {}
            current_course = None
            
            for _, row in df.iterrows():
                course_code = row['course_code']
                
                # Skip empty rows
                if not course_code:
                    continue
                
                # Create hour distribution
                hours = HourDistribution(
                    lecture=row['Lecture'],
                    lab=row['Lab'],
                    practice=row['Practice'],
                    seminar=row['Seminar'],
                    individual=row['Individual']
                )
                
                # Create semester data
                semester_data = SemesterData(
                    semester=row['semester'],
                    credits=row['credits'],
                    hours=hours
                ).to_dict()
                
                if course_code not in curriculum_data:
                    # New course
                    curriculum_data[course_code] = {
                        'course_code': course_code,
                        'course_name': row['course_name'],
                        'credits': row['credits'],
                        'semester': row['semester'],
                        'prerequisites': self.process_prerequisites(row.get('prerequisites', ''))
                    }
                else:
                    # Additional semester for existing course
                    curriculum_data[course_code]['semesters'] = [semester_data]
            
            # Validate the constructed data
            CurriculumSchema.validate_curriculum(curriculum_data)
            return curriculum_data
            
        except Exception as e:
            raise ValidationError(f"Error processing Excel file: {str(e)}")

    def export_excel(self, curriculum_data: Dict, file_path: str) -> None:
        """Export curriculum data to Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Curriculum"

        # Write headers
        headers = self.REQUIRED_COLUMNS + ['Prerequisites']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Write data
        row_num = 2
        for course_code, course_data in curriculum_data.items():
            first_semester = True
            for semester in course_data['semesters']:
                row_data = [
                    course_code if first_semester else '',
                    course_data['course_name'] if first_semester else '',
                    semester['credits'],
                    semester['semester'],
                    ', '.join(course_data['prerequisites']) if first_semester else ''
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.value = value
                    
                    # Apply styling for merged course data
                    if not first_semester and col <= 3:
                        cell.fill = PatternFill(start_color="EEEEEE", 
                                              end_color="EEEEEE", 
                                              fill_type="solid")
                
                first_semester = False
                row_num += 1

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True

        # Save the workbook
        wb.save(file_path)

    def preview_data(self, file_path: str) -> Tuple[Dict, List[str]]:
        """Preview data from Excel file and return any warnings"""
        warnings = []
        try:
            data = self.read_excel(file_path)
            
            # Generate warnings for potential issues
            for code, course in data.items():
                # Check for courses with high credit loads
                total_credits = sum(sem['credits'] for sem in course['semesters'])
                if total_credits > 8:
                    warnings.append(
                        f"Warning: Course {code} has unusually high credits ({total_credits})"
                    )
                
                # Check for courses with many prerequisites
                if len(course['prerequisites']) > 3:
                    warnings.append(
                        f"Warning: Course {code} has many prerequisites ({len(course['prerequisites'])})"
                    )
                
                # Check for courses spanning many semesters
                if len(course['semesters']) > 2:
                    warnings.append(
                        f"Warning: Course {code} spans {len(course['semesters'])} semesters"
                    )
            
            return data, warnings
            
        except ValidationError as e:
            raise ValidationError(f"Validation error during preview: {str(e)}")
        except Exception as e:
            raise ValidationError(f"Error during preview: {str(e)}")

    @staticmethod
    def generate_template(file_path: str) -> None:
        """Generate an empty template Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Curriculum Template"
        
        # Headers
        headers = [
            'course_code',
            'course_name',
            'credits',
            'semester',
            'prerequisites'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", 
                                  end_color="CCCCCC", 
                                  fill_type="solid")
        
        # Add example data
        example_data = [
            ['CS101', 'Introduction to Programming', 3, 1, 'CS100, MATH101'],
            ['CS201', 'Data Structures', 3, 2, 'CS101'],
            ['', '', 2, 3, '']  # Multi-semester example
        ]
        
        for row_num, row_data in enumerate(example_data, 2):
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = value
                if row_num == 4 and col <= 3:  # Style continued course rows
                    cell.fill = PatternFill(start_color="EEEEEE", 
                                          end_color="EEEEEE", 
                                          fill_type="solid")
        
        # Add validation and notes
        ws['M1'] = "Notes:"
        notes = [
            "- course_code: Unique identifier for each course",
            "- credits × 30 must equal total hours",
            "- prerequisites: Comma-separated course codes",
            "- For multi-semester courses, leave course_code/course_name/credits empty in continuation rows"
        ]
        
        for i, note in enumerate(notes, 2):
            ws[f'M{i}'] = note
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 3):
            ws.column_dimensions[get_column_letter(col)].auto_size = True
        
        wb.save(file_path) 